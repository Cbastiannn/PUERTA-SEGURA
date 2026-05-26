import io, csv, json
from datetime import timedelta, date

from django.contrib.auth import authenticate, get_user_model
from django.http import HttpResponse
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Count, Q, Prefetch
from django.db.models.functions import ExtractHour, TruncDate
from django.conf import settings

from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken

from django_filters import rest_framework as df_filters

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (Vehiculo, RegistroAcceso, ListaNegra, Visitante,
                     EstadoBarrera, Auditoria, Notificacion, validate_colombian_plate)
from .serializers import (UserSerializer, UserCreateSerializer, UserUpdateSerializer,
                           VehiculoSerializer, RegistroAccesoSerializer,
                           ListaNegraSerializer, VisitanteSerializer,
                           EstadoBarreraSerializer, AuditoriaSerializer,
                           NotificacionSerializer)

User = get_user_model()


# ── Helpers ───────────────────────────────────────────────────────────────────
def get_client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    return xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR')


def broadcast_stats():
    """Envía stats actualizadas por WebSocket a todos los clientes del dashboard."""
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        hoy = timezone.now().date()
        acc = RegistroAcceso.objects.filter(timestamp__date=hoy)
        data = {
            'vehiculos_activos': Vehiculo.objects.filter(activo=True).count(),
            'accesos_hoy':       acc.count(),
            'autorizados_hoy':   acc.filter(estado='Autorizado').count(),
            'denegados_hoy':     acc.filter(estado='Denegado').count(),
            'en_lista_negra':    ListaNegra.objects.filter(activo=True).count(),
            'visitantes_activos':Visitante.objects.filter(hora_salida__isnull=True).count(),
            'timestamp':         timezone.now().isoformat(),
        }
        cache.delete('dashboard_stats')  # invalidar caché
        cl = get_channel_layer()
        async_to_sync(cl.group_send)('dashboard', {'type': 'stats_update', 'data': data})
    except Exception:
        pass  # Redis no disponible — degradar gracefully


def broadcast_barrier(barrier_obj):
    """Envía nuevo estado de barrera por WebSocket."""
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        data = {'abierta': barrier_obj.abierta, 'razon': barrier_obj.razon,
                'actualizado_en': barrier_obj.actualizado_en.isoformat()}
        cl = get_channel_layer()
        async_to_sync(cl.group_send)('barrier', {'type': 'barrier_update', 'data': data})
        async_to_sync(cl.group_send)('dashboard', {'type': 'barrier_update', 'data': data})
    except Exception:
        pass


def notify_user_ws(user_id, notif_data):
    """Envía notificación push por WebSocket al usuario específico."""
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        cl = get_channel_layer()
        async_to_sync(cl.group_send)(
            f'notifications_{user_id}',
            {'type': 'notification', 'data': notif_data}
        )
    except Exception:
        pass


def _validar_acceso(plate):
    if ListaNegra.objects.filter(placa=plate, activo=True).exists():
        bl = ListaNegra.objects.get(placa=plate, activo=True)
        # Notificar admins
        for admin in User.objects.filter(role='vigilante', is_active=True):
            n = Notificacion.objects.create(
                usuario=admin, tipo='alerta',
                titulo=f'🚫 Lista Negra — {plate}',
                mensaje=f'Intentó ingresar. Motivo: {bl.motivo}',
                placa=plate,
            )
            notify_user_ws(admin.id, {'id': n.id, 'tipo': 'alerta',
                                       'titulo': n.titulo, 'mensaje': n.mensaje, 'placa': plate})
        return False, 'Denegado', f'LISTA NEGRA: {bl.motivo}', None

    try:
        v = Vehiculo.objects.get(placa=plate)
    except Vehiculo.DoesNotExist:
        return False, 'Denegado', 'Placa no registrada en el sistema', None

    if not v.activo:
        return False, 'Denegado', 'Vehículo DESACTIVADO por administrador', v
    if not v.qr_vigente():
        return False, 'Denegado', f'QR vencido el {v.expira_en.strftime("%d/%m/%Y")}', v
    return True, 'Autorizado', '', v


# ── Auth ──────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    email    = request.data.get('email', '').strip()
    password = request.data.get('password', '')
    ip       = get_client_ip(request)

    try:
        user_obj = User.objects.get(email=email)
    except User.DoesNotExist:
        Auditoria.log(None, 'login_failed', f'Email desconocido: {email}', ip=ip, exitoso=False)
        return Response({'error': 'Credenciales incorrectas'}, status=400)

    if user_obj.is_locked():
        return Response({'error': 'Cuenta bloqueada.', 'locked': True,
                         'seconds': user_obj.remaining_lock_seconds()}, status=403)
    if not user_obj.is_active:
        return Response({'error': 'Usuario inactivo.'}, status=403)

    user = authenticate(request, username=user_obj.username, password=password)
    if not user:
        user_obj.register_failed_attempt()
        Auditoria.log(user_obj, 'login_failed', f'Intento {user_obj.login_attempts}/5', ip=ip, exitoso=False)
        return Response({'error': f'Contraseña incorrecta. Restantes: {max(0, 5 - user_obj.login_attempts)}',
                         'attempts': user_obj.login_attempts}, status=400)

    user.reset_attempts()
    Auditoria.log(user, 'login', f'IP: {ip}', ip=ip)
    refresh = RefreshToken.for_user(user)

    response = Response({'user': UserSerializer(user).data,
                         'token': str(refresh.access_token),  # también en body para el cliente
                         'refresh': str(refresh)})

    # Setear httpOnly cookies
    jwt_settings = settings.SIMPLE_JWT
    response.set_cookie(
        jwt_settings.get('AUTH_COOKIE', 'access_token'),
        str(refresh.access_token),
        httponly=jwt_settings.get('AUTH_COOKIE_HTTP_ONLY', True),
        secure=jwt_settings.get('AUTH_COOKIE_SECURE', False),
        samesite=jwt_settings.get('AUTH_COOKIE_SAMESITE', 'Lax'),
        max_age=int(jwt_settings['ACCESS_TOKEN_LIFETIME'].total_seconds()),
    )
    response.set_cookie(
        jwt_settings.get('AUTH_COOKIE_REFRESH', 'refresh_token'),
        str(refresh),
        httponly=True,
        secure=jwt_settings.get('AUTH_COOKIE_SECURE', False),
        samesite=jwt_settings.get('AUTH_COOKIE_SAMESITE', 'Lax'),
        max_age=int(jwt_settings['REFRESH_TOKEN_LIFETIME'].total_seconds()),
    )
    return response


@api_view(['POST'])
def logout(request):
    try:
        refresh_token = (request.data.get('refresh') or
                         request.COOKIES.get(settings.SIMPLE_JWT.get('AUTH_COOKIE_REFRESH', 'refresh_token')))
        if refresh_token:
            RefreshToken(refresh_token).blacklist()
    except Exception:
        pass
    Auditoria.log(request.user, 'logout', '', ip=get_client_ip(request))
    response = Response({'message': 'Sesión cerrada'})
    response.delete_cookie(settings.SIMPLE_JWT.get('AUTH_COOKIE', 'access_token'))
    response.delete_cookie(settings.SIMPLE_JWT.get('AUTH_COOKIE_REFRESH', 'refresh_token'))
    return response


@api_view(['GET'])
def me(request):
    return Response(UserSerializer(request.user).data)


# ── Usuarios ──────────────────────────────────────────────────────────────────
class UserViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    search_fields      = ['first_name', 'last_name', 'email']
    ordering           = ['first_name']

    def get_queryset(self):
        return User.objects.all()

    def get_serializer_class(self):
        if self.action == 'create': return UserCreateSerializer
        if self.action in ['update', 'partial_update']: return UserUpdateSerializer
        return UserSerializer

    def _admin_only(self):
        if self.request.user.role != 'vigilante':
            return Response({'error': 'Solo administradores'}, status=403)

    def create(self, request, *args, **kwargs):
        err = self._admin_only()
        if err: return err
        resp = super().create(request, *args, **kwargs)
        if resp.status_code == 201:
            Auditoria.log(request.user, 'usuario_creado',
                          f"Nuevo: {request.data.get('email')}", ip=get_client_ip(request))
        return resp

    def update(self, request, *args, **kwargs):
        err = self._admin_only()
        if err: return err
        return super().update(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        user = self.get_object()
        pw = request.data.get('password', '')
        if len(pw) < 8: return Response({'error': 'Mínimo 8 caracteres'}, status=400)
        user.set_password(pw); user.save()
        Auditoria.log(request.user, 'usuario_modificado', f'Reset pwd: {user.email}', ip=get_client_ip(request))
        return Response({'message': 'Contraseña actualizada'})

    @action(detail=True, methods=['patch'])
    def toggle_active(self, request, pk=None):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        user = self.get_object()
        if user.id == request.user.id:
            return Response({'error': 'No puedes desactivarte a ti mismo'}, status=400)
        user.is_active = not user.is_active; user.save()
        return Response(UserSerializer(user).data)


# ── Vehículos ─────────────────────────────────────────────────────────────────
class VehiculoFilter(df_filters.FilterSet):
    placa        = df_filters.CharFilter(lookup_expr='icontains')
    propietario  = df_filters.CharFilter(lookup_expr='icontains')
    tipo         = df_filters.ChoiceFilter(choices=Vehiculo.TIPOS)
    activo       = df_filters.BooleanFilter()
    expira_antes = df_filters.DateFilter(field_name='expira_en', lookup_expr='lte')
    expira_despues = df_filters.DateFilter(field_name='expira_en', lookup_expr='gte')
    class Meta:
        model = Vehiculo
        fields = ['placa', 'propietario', 'tipo', 'activo']


class VehiculoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class   = VehiculoSerializer
    filterset_class    = VehiculoFilter
    search_fields      = ['placa', 'propietario', 'marca', 'modelo', 'documento']
    ordering_fields    = ['placa', 'propietario', 'creado_en', 'expira_en', 'tipo']
    ordering           = ['-creado_en']

    def get_queryset(self):
        return Vehiculo.objects.select_related('registrado_por')

def create(self, request, *args, **kwargs):
    if request.user.role != 'vigilante':
        return Response({'error': 'Solo administradores'}, status=403)
    resp = super().create(request, *args, **kwargs)
    if resp.status_code == 201:
        v = Vehiculo.objects.get(id=resp.data['id'])
        Auditoria.log(request.user, 'vehiculo_creado', v.placa, ip=get_client_ip(request))
    return resp

    def perform_create(self, serializer):
        serializer.save(registrado_por=self.request.user)

    @action(detail=True, methods=['patch'])
    def toggle(self, request, pk=None):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        v = self.get_object(); v.activo = not v.activo; v.save()
        Auditoria.log(request.user, 'vehiculo_modificado',
                      f'{v.placa} {"activado" if v.activo else "desactivado"}', ip=get_client_ip(request))
        return Response(VehiculoSerializer(v).data)

    @action(detail=True, methods=['get'])
    def qr(self, request, pk=None):
        v = self.get_object()
        return Response({
            'placa': v.placa, 'propietario': v.propietario, 'tipo': v.tipo,
            'marca': v.marca, 'modelo': v.modelo, 'color': v.color,
            'codigo_qr': v.codigo_qr, 'dias_vigencia': v.dias_vigencia(),
            'activo': v.activo, 'expira_en': v.expira_en,
            'qr_imagen_url': v.qr_imagen_url(), 'en_lista_negra': v.en_lista_negra,
        })

    @action(detail=True, methods=['post'])
    def renovar_qr(self, request, pk=None):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        v = self.get_object()
        from .models import generate_qr_token
        v.codigo_qr = generate_qr_token(v.placa)
        v.expira_en = timezone.now() + timedelta(days=settings.QR_VALIDITY_DAYS)
        v.save()
        from .tasks import enviar_qr_por_correo
        enviar_qr_por_correo.delay(v.id)
        Auditoria.log(request.user, 'vehiculo_modificado',
                      f'QR renovado: {v.placa}', ip=get_client_ip(request))
        return Response({'message': 'QR renovado y enviado por correo',
                         'expira_en': v.expira_en, 'dias_vigencia': v.dias_vigencia()})

    @action(detail=True, methods=['post'])
    def reenviar_qr(self, request, pk=None):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        v = self.get_object()
        if not v.correo: return Response({'error': 'Sin correo registrado'}, status=400)
        from .tasks import enviar_qr_por_correo
        enviar_qr_por_correo.delay(v.id)
        return Response({'message': f'QR en cola de envío → {v.correo}'})

    @action(detail=False, methods=['post'])
    def importar_excel(self, request):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        file = request.FILES.get('file')
        if not file: return Response({'error': 'Adjunta un archivo .xlsx'}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'error': 'Archivo Excel inválido'}, status=400)

        creados, errores = [], []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]: continue
            try:
                cols = [(str(row[j]) if j < len(row) and row[j] else '') for j in range(9)]
                placa_raw, tipo, marca, modelo, color, propietario, documento, correo, telefono = cols
                plate, tipo_det = validate_colombian_plate(placa_raw)
                if Vehiculo.objects.filter(placa=plate).exists():
                    errores.append(f'Fila {i}: {plate} ya existe'); continue
                v = Vehiculo.objects.create(
                    placa=plate, tipo=tipo or tipo_det, marca=marca, modelo=modelo,
                    color=color, propietario=propietario, documento=documento,
                    correo=correo, telefono=telefono, registrado_por=request.user,
                )
                creados.append(plate)
                from .tasks import enviar_qr_por_correo
                if v.correo: enviar_qr_por_correo.delay(v.id)
            except Exception as e:
                errores.append(f'Fila {i}: {e}')

        Auditoria.log(request.user, 'vehiculo_creado',
                      f'Importación: {len(creados)} creados, {len(errores)} errores', ip=get_client_ip(request))
        return Response({'creados': len(creados), 'placas': creados, 'errores': errores})

    @action(detail=False, methods=['get'])
    def template_excel(self, request):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Plantilla'
        headers = ['Placa*', 'Tipo', 'Marca', 'Modelo', 'Color', 'Propietario*', 'Documento', 'Correo', 'Teléfono']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill(start_color='050A14', end_color='050A14', fill_type='solid')
        ws.append(['ABC123', 'Automóvil', 'Toyota', 'Corolla', 'Blanco', 'Juan García', '12345678', 'juan@correo.com', '3001234567'])
        ws.append(['XYZ12D', 'Motocicleta', 'Yamaha', 'FZ', 'Rojo', 'María Pérez', '98765432', 'maria@correo.com', '3109876543'])
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="plantilla-vehiculos.xlsx"'
        return resp

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Vehículos'
        ws.merge_cells('A1:J1'); ws['A1'] = 'PUERTA SEGURA v3.0 — Vehículos'
        ws['A1'].font = Font(bold=True, size=13, color='00B4FF')
        for col, h in enumerate(['Placa','Propietario','Tipo','Marca','Modelo','Color','Correo','Teléfono','Estado','Vence'], 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill(start_color='0A1520', end_color='0A1520', fill_type='solid')
        for ri, v in enumerate(qs, 4):
            ws.cell(ri,1,v.placa).font = Font(color='00B4FF', bold=True)
            for ci, val in enumerate([v.propietario,v.tipo,v.marca,v.modelo,v.color,v.correo,v.telefono,'Activo' if v.activo else 'Inactivo',v.expira_en.strftime('%d/%m/%Y')],2):
                ws.cell(ri, ci, val)
        for col, w in enumerate([12,28,14,14,14,12,28,14,12,12],1):
            ws.column_dimensions[get_column_letter(col)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="vehiculos.xlsx"'
        return resp


# ── Acceso ────────────────────────────────────────────────────────────────────
@api_view(['POST'])
def escanear_qr(request):
    qr_data = request.data.get('qr_code', '').strip()
    if not qr_data.startswith('PSU-v2-'):
        return Response({'error': 'QR no pertenece a Puerta Segura'}, status=400)
    parts = qr_data.split('-')
    if len(parts) < 9: return Response({'error': 'QR inválido'}, status=400)
    plate = parts[7]
    ok, estado, motivo, v = _validar_acceso(plate)
    ip = get_client_ip(request)
    RegistroAcceso.objects.create(
        vehiculo=v, placa=plate, propietario=v.propietario if v else 'Desconocido',
        movimiento='Entrada' if ok else 'Intento', metodo='QR',
        punto_acceso=request.user.access_point, estado=estado,
        motivo_denegacion=motivo, registrado_por=request.user, ip_address=ip,
    )
    if ok:
        b = EstadoBarrera.get_estado(); b.abierta = True
        b.actualizado_por = request.user; b.razon = f'QR: {plate}'; b.save()
        broadcast_barrier(b)
        Auditoria.log(request.user, 'acceso_autorizado', f'QR: {plate}', ip=ip)
    else:
        Auditoria.log(request.user, 'acceso_denegado', f'QR: {plate}', ip=ip, exitoso=False)
    broadcast_stats()
    return Response({'autorizado': ok, 'estado': estado, 'motivo': motivo, 'placa': plate,
                     'propietario': v.propietario if v else None,
                     'vehiculo': VehiculoSerializer(v).data if v else None})


@api_view(['POST'])
def verificar_placa(request):
    placa_raw = request.data.get('placa', '').strip()
    ip = get_client_ip(request)
    try:
        plate, tipo = validate_colombian_plate(placa_raw)
    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    ok, estado, motivo, v = _validar_acceso(plate)
    RegistroAcceso.objects.create(
        vehiculo=v, placa=plate, propietario=v.propietario if v else 'Desconocido',
        movimiento='Entrada' if ok else 'Intento', metodo='Manual',
        punto_acceso=request.user.access_point, estado=estado,
        motivo_denegacion=motivo, registrado_por=request.user, ip_address=ip,
    )
    if ok: Auditoria.log(request.user, 'acceso_autorizado', f'Manual: {plate}', ip=ip)
    else:  Auditoria.log(request.user, 'acceso_denegado', f'Manual: {plate}', ip=ip, exitoso=False)
    broadcast_stats()
    return Response({'autorizado': ok, 'estado': estado, 'motivo': motivo, 'placa': plate,
                     'tipo_vehiculo': tipo, 'vehiculo': VehiculoSerializer(v).data if v else None})


@api_view(['POST'])
def registrar_salida(request):
    try:
        plate, _ = validate_colombian_plate(request.data.get('placa', '').strip())
    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    v = Vehiculo.objects.filter(placa=plate).first()
    r = RegistroAcceso.objects.create(
        vehiculo=v, placa=plate, propietario=v.propietario if v else 'Desconocido',
        movimiento='Salida', metodo=request.data.get('metodo', 'Manual'),
        punto_acceso=request.user.access_point, estado='Autorizado',
        registrado_por=request.user, ip_address=get_client_ip(request),
    )
    broadcast_stats()
    return Response({'message': f'Salida registrada: {plate}', 'registro_id': r.id})


@api_view(['GET', 'POST'])
def barrera(request):
    b = EstadoBarrera.get_estado()
    if request.method == 'POST':
        b.abierta = request.data.get('abierta', False)
        b.actualizado_por = request.user
        b.razon = request.data.get('razon', 'Control manual')
        b.save()
        broadcast_barrier(b)
        accion = 'barrera_abierta' if b.abierta else 'barrera_cerrada'
        Auditoria.log(request.user, accion, b.razon, ip=get_client_ip(request))
    return Response(EstadoBarreraSerializer(b).data)


# ── Historial ─────────────────────────────────────────────────────────────────
class HistorialFilter(df_filters.FilterSet):
    placa        = df_filters.CharFilter(lookup_expr='icontains')
    estado       = df_filters.ChoiceFilter(choices=RegistroAcceso.ESTADOS)
    movimiento   = df_filters.ChoiceFilter(choices=RegistroAcceso.MOVIMIENTOS)
    metodo       = df_filters.ChoiceFilter(choices=RegistroAcceso.METODOS)
    desde        = df_filters.DateTimeFilter(field_name='timestamp', lookup_expr='gte')
    hasta        = df_filters.DateTimeFilter(field_name='timestamp', lookup_expr='lte')
    punto_acceso = df_filters.CharFilter(lookup_expr='icontains')
    class Meta:
        model  = RegistroAcceso
        fields = ['placa', 'estado', 'movimiento', 'metodo', 'punto_acceso']


class HistorialViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class   = RegistroAccesoSerializer
    filterset_class    = HistorialFilter
    search_fields      = ['placa', 'propietario', 'punto_acceso']
    ordering_fields    = ['timestamp', 'placa', 'estado', 'movimiento']
    ordering           = ['-timestamp']

    def get_queryset(self):
        return RegistroAcceso.objects.select_related('registrado_por')

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        qs = self.filter_queryset(self.get_queryset())[:2000]
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Historial'
        ws.merge_cells('A1:H1'); ws['A1'] = 'PUERTA SEGURA v3.0 — Historial'
        ws['A1'].font = Font(bold=True, size=13, color='00B4FF')
        for col, h in enumerate(['#','Placa','Propietario','Movimiento','Método','Punto','Fecha y Hora','Estado'],1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill(start_color='0A1520', end_color='0A1520', fill_type='solid')
        for ri, a in enumerate(qs, 4):
            ws.cell(ri,1,ri-3); ws.cell(ri,2,a.placa).font=Font(color='00B4FF',bold=True)
            ws.cell(ri,3,a.propietario); ws.cell(ri,4,a.movimiento); ws.cell(ri,5,a.metodo)
            ws.cell(ri,6,a.punto_acceso); ws.cell(ri,7,a.timestamp.strftime('%d/%m/%Y %H:%M'))
            c=ws.cell(ri,8,a.estado); c.font=Font(color='00FFA3' if a.estado=='Autorizado' else 'FF4444',bold=True)
        for col,w in enumerate([5,12,28,14,10,22,20,12],1):
            ws.column_dimensions[get_column_letter(col)].width=w
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        resp=HttpResponse(buf.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition']='attachment; filename="historial.xlsx"'
        return resp

    @action(detail=False, methods=['get'])
    def exportar_csv(self, request):
        qs = self.filter_queryset(self.get_queryset())[:5000]
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="historial.csv"'
        resp.write('\ufeff')
        w = csv.writer(resp)
        w.writerow(['Placa','Propietario','Movimiento','Método','Punto','Fecha','Estado','Motivo'])
        for a in qs:
            w.writerow([a.placa,a.propietario,a.movimiento,a.metodo,a.punto_acceso,
                        a.timestamp.strftime('%d/%m/%Y %H:%M'),a.estado,a.motivo_denegacion])
        return resp


# ── Lista Negra, Visitantes, Notificaciones, Auditoría ───────────────────────
class ListaNegraViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class   = ListaNegraSerializer
    queryset           = ListaNegra.objects.filter(activo=True).order_by('-agregado_en')
    search_fields      = ['placa', 'motivo']

    def create(self, request, *args, **kwargs):
        if request.user.role != 'vigilante': return Response({'error': 'Solo admins'}, status=403)
        resp = super().create(request, *args, **kwargs)
        if resp.status_code == 201:
            Auditoria.log(request.user, 'lista_negra_agregado',
                          f"Placa {request.data.get('placa')}", ip=get_client_ip(request))
        return resp

    def perform_create(self, serializer):
        serializer.save(agregado_por=self.request.user)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'vigilante': return Response({'error': 'Sin permiso'}, status=403)
        placa = self.get_object().placa
        resp = super().destroy(request, *args, **kwargs)
        Auditoria.log(request.user, 'lista_negra_removido', f'Placa {placa}', ip=get_client_ip(request))
        return resp


class VisitanteViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class   = VisitanteSerializer
    queryset           = Visitante.objects.all()
    search_fields      = ['nombre', 'placa', 'anfitrion', 'documento']
    ordering           = ['-hora_entrada']

    def perform_create(self, serializer):
        serializer.save(registrado_por=self.request.user)

    @action(detail=True, methods=['patch'])
    def salida(self, request, pk=None):
        v = self.get_object()
        if v.hora_salida: return Response({'error': 'Salida ya registrada'}, status=400)
        v.hora_salida = timezone.now()
        v.observaciones = request.data.get('observaciones', v.observaciones)
        v.save()
        return Response(VisitanteSerializer(v).data)


class NotificacionViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class   = NotificacionSerializer
    ordering           = ['-creada_en']

    def get_queryset(self):
        return Notificacion.objects.filter(usuario=self.request.user)

    @action(detail=False, methods=['post'])
    def marcar_todas_leidas(self, request):
        Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
        return Response({'message': 'Listo'})

    @action(detail=True, methods=['patch'])
    def leer(self, request, pk=None):
        n = self.get_object(); n.leida = True; n.save()
        return Response(NotificacionSerializer(n).data)


class AuditoriaViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class   = AuditoriaSerializer
    ordering           = ['-timestamp']

    def get_queryset(self):
        qs = Auditoria.objects.all()
        if self.request.user.role != 'vigilante':
            qs = qs.filter(usuario=self.request.user)
        return qs


# ── Stats (con caché Redis) ───────────────────────────────────────────────────
@api_view(['GET'])
def stats(request):
    CACHE_KEY = f'dashboard_stats_{request.user.id}'
    cached = cache.get(CACHE_KEY)
    if cached:
        return Response(cached)

    hoy  = timezone.now().date()
    ayer = hoy - timedelta(days=1)
    acc_hoy  = RegistroAcceso.objects.filter(timestamp__date=hoy)
    acc_ayer = RegistroAcceso.objects.filter(timestamp__date=ayer)

    por_hora = list(
        RegistroAcceso.objects.filter(timestamp__date=hoy)
        .annotate(hora=ExtractHour('timestamp'))
        .values('hora').annotate(total=Count('id')).order_by('hora')
    )

    # Vehículos en campus — query optimizada con una sola pasada
    en_campus = []
    visto = set()
    entradas = (RegistroAcceso.objects
                .filter(movimiento='Entrada', estado='Autorizado')
                .order_by('-timestamp')
                .values('placa', 'propietario', 'timestamp'))
    for e in entradas:
        if e['placa'] in visto: continue
        visto.add(e['placa'])
        salida = RegistroAcceso.objects.filter(
            placa=e['placa'], movimiento='Salida', timestamp__gt=e['timestamp']
        ).exists()
        if not salida:
            h = (timezone.now() - e['timestamp']).total_seconds() / 3600
            en_campus.append({'placa': e['placa'], 'propietario': e['propietario'],
                               'desde': e['timestamp'].isoformat(), 'horas': round(h, 1), 'alerta': h > 4})
        if len(en_campus) >= 100: break

    hoy_total  = acc_hoy.count()
    ayer_total = acc_ayer.count()
    cambio_pct = round(((hoy_total - ayer_total) / max(ayer_total, 1)) * 100, 1)

    por_dia = []
    for i in range(13, -1, -1):
        d = hoy - timedelta(days=i)
        por_dia.append({'dia': d.strftime('%d/%m'), 'total': RegistroAcceso.objects.filter(timestamp__date=d).count()})

    por_tipo = list(Vehiculo.objects.values('tipo').annotate(total=Count('id')).order_by('-total'))
    top_placas = list(
        RegistroAcceso.objects.filter(timestamp__date__gte=hoy - timedelta(days=30))
        .values('placa', 'propietario').annotate(total=Count('id')).order_by('-total')[:5]
    )

    data = {
        'vehiculos_totales':         Vehiculo.objects.count(),
        'vehiculos_activos':         Vehiculo.objects.filter(activo=True).count(),
        'vehiculos_proximos_vencer': Vehiculo.objects.filter(
            activo=True, expira_en__lte=timezone.now() + timedelta(days=30)).count(),
        'accesos_hoy':       hoy_total,
        'autorizados_hoy':   acc_hoy.filter(estado='Autorizado').count(),
        'denegados_hoy':     acc_hoy.filter(estado='Denegado').count(),
        'accesos_ayer':      ayer_total,
        'cambio_porcentaje': cambio_pct,
        'en_campus':         len(en_campus),
        'visitantes_activos':Visitante.objects.filter(hora_salida__isnull=True).count(),
        'en_lista_negra':    ListaNegra.objects.filter(activo=True).count(),
        'total_usuarios':    User.objects.filter(is_active=True).count(),
        'notificaciones_no_leidas': Notificacion.objects.filter(usuario=request.user, leida=False).count(),
        'alertas_permanencia': [v for v in en_campus if v['alerta']],
        'vehiculos_en_campus': en_campus,
        'por_hora':  por_hora,
        'por_dia':   por_dia,
        'por_tipo':  por_tipo,
        'top_placas':top_placas,
    }
    cache.set(CACHE_KEY, data, settings.STATS_CACHE_SECONDS)
    return Response(data)


@api_view(['GET'])
def reporte_resumen(request):
    desde_str = request.query_params.get('desde', '')
    hasta_str = request.query_params.get('hasta', '')
    try:
        desde = date.fromisoformat(desde_str) if desde_str else date.today() - timedelta(days=30)
        hasta = date.fromisoformat(hasta_str) if hasta_str else date.today()
    except ValueError:
        return Response({'error': 'Formato de fecha inválido (YYYY-MM-DD)'}, status=400)

    qs = RegistroAcceso.objects.filter(timestamp__date__gte=desde, timestamp__date__lte=hasta)
    return Response({
        'periodo':            {'desde': desde.isoformat(), 'hasta': hasta.isoformat()},
        'total_accesos':      qs.count(),
        'autorizados':        qs.filter(estado='Autorizado').count(),
        'denegados':          qs.filter(estado='Denegado').count(),
        'por_metodo':         list(qs.values('metodo').annotate(total=Count('id'))),
        'por_dia':            list(qs.annotate(dia=TruncDate('timestamp')).values('dia').annotate(total=Count('id')).order_by('dia')),
        'vehiculos_unicos':   qs.values('placa').distinct().count(),
        'nuevos_vehiculos':   Vehiculo.objects.filter(creado_en__date__gte=desde, creado_en__date__lte=hasta).count(),
        'nuevos_visitantes':  Visitante.objects.filter(hora_entrada__date__gte=desde, hora_entrada__date__lte=hasta).count(),
    })


# ── Exportación PDF en servidor (sin límite de memoria del navegador) ─────────
@api_view(['GET'])
def exportar_historial_pdf(request):
    """Genera PDF del historial en el servidor usando reportlab o texto plano."""
    qs_filter = HistorialFilter(request.GET, queryset=RegistroAcceso.objects.all())
    qs = qs_filter.qs.order_by('-timestamp')[:5000]

    # Generamos un CSV estilizado como fallback si reportlab no está
    # En producción se puede agregar: pip install reportlab
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        buf = io.BytesIO()
        c_pdf = canvas.Canvas(buf, pagesize=landscape(A4))
        w, h = landscape(A4)
        c_pdf.setFillColorRGB(0.02, 0.04, 0.06)
        c_pdf.rect(0, 0, w, h, fill=True)
        c_pdf.setFillColorRGB(0, 0.71, 1)
        c_pdf.setFont('Helvetica-Bold', 14)
        c_pdf.drawString(30, h-40, 'PUERTA SEGURA v3.0 — Historial de Accesos')
        c_pdf.setFillColorRGB(0.35, 0.49, 0.6)
        c_pdf.setFont('Helvetica', 9)
        c_pdf.drawString(30, h-58, f'Universidad de Cundinamarca · {timezone.now().strftime("%d/%m/%Y %H:%M")} · {qs.count()} registros')
        headers = ['Placa','Propietario','Movimiento','Método','Punto de Acceso','Fecha y Hora','Estado']
        xs = [30,90,220,295,345,460,590]
        y = h - 85
        c_pdf.setFont('Helvetica-Bold', 8); c_pdf.setFillColorRGB(0, 0.71, 1)
        for i, header in enumerate(headers): c_pdf.drawString(xs[i], y, header)
        y -= 4; c_pdf.setStrokeColorRGB(0, 0.71, 1); c_pdf.setLineWidth(0.5); c_pdf.line(30, y, w-30, y)
        y -= 12; c_pdf.setFont('Helvetica', 7.5)
        for acc in qs:
            if y < 40: c_pdf.showPage(); c_pdf.setFillColorRGB(0.02,0.04,0.06); c_pdf.rect(0,0,w,h,fill=True); y = h-30; c_pdf.setFont('Helvetica',7.5)
            auth = acc.estado == 'Autorizado'
            c_pdf.setFillColorRGB(0,0.78,0.64) if auth else c_pdf.setFillColorRGB(1,0.31,0.31)
            vals = [acc.placa, acc.propietario[:20], acc.movimiento, acc.metodo, (acc.punto_acceso or '')[:16], acc.timestamp.strftime('%d/%m/%Y %H:%M'), acc.estado]
            for i, val in enumerate(vals): c_pdf.drawString(xs[i], y, str(val))
            y -= 10
        c_pdf.save(); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="historial-puerta-segura.pdf"'
        return resp
    except ImportError:
        # Fallback: CSV si no hay reportlab
        return Response({'error': 'reportlab no instalado. Usa el export PDF desde el navegador.'}, status=501)


# ── Exportación de auditoría ──────────────────────────────────────────────────
@api_view(['GET'])
def exportar_auditoria_excel(request):
    if request.user.role != 'vigilante':
        return Response({'error': 'Solo administradores'}, status=403)
    qs = Auditoria.objects.all().select_related('usuario').order_by('-timestamp')[:5000]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Auditoría'
    ws.merge_cells('A1:F1'); ws['A1'] = 'PUERTA SEGURA v3.0 — Bitácora de Auditoría'
    ws['A1'].font = Font(bold=True, size=13, color='00B4FF')
    for col, h in enumerate(['Fecha y Hora','Usuario','Acción','Detalle','IP','Resultado'], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='0A1520', end_color='0A1520', fill_type='solid')
    for ri, a in enumerate(qs, 4):
        ws.cell(ri,1, a.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(ri,2, a.usuario.get_full_name() if a.usuario else 'Sistema')
        ws.cell(ri,3, a.accion.replace('_',' '))
        ws.cell(ri,4, a.detalle[:120])
        ws.cell(ri,5, a.ip_address or '')
        c = ws.cell(ri,6, 'OK' if a.exitoso else 'FALLO')
        c.font = Font(color='00FFA3' if a.exitoso else 'FF4444', bold=True)
    for col, w in enumerate([18,24,22,44,14,10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="auditoria-puerta-segura.xlsx"'
    return resp


# ── Estadísticas avanzadas: heatmap ──────────────────────────────────────────
@api_view(['GET'])
def heatmap_data(request):
    """
    Devuelve datos para la gráfica de calor (hora × día de semana).
    Últimos 90 días de accesos, agrupados por día_semana y hora.
    """
    desde = timezone.now() - timedelta(days=90)
    data = (
        RegistroAcceso.objects
        .filter(timestamp__gte=desde, estado='Autorizado')
        .annotate(
            hora=ExtractHour('timestamp'),
            dia=TruncDate('timestamp'),
        )
        .values('hora', 'dia')
        .annotate(total=Count('id'))
        .order_by()
    )

    # Construir matriz hora (0-23) × día_semana (0=lunes..6=domingo)
    from django.db.models.functions import ExtractWeekDay
    matrix = {}
    for row in data:
        h   = row['hora']
        dow = row['dia'].weekday()  # 0=lunes, 6=domingo
        key = f"{dow}-{h}"
        matrix[key] = matrix.get(key, 0) + row['total']

    result = [
        {'dia': dow, 'hora': h, 'total': matrix.get(f"{dow}-{h}", 0)}
        for dow in range(7)
        for h in range(6, 22)  # solo horas laborales 6am-10pm
    ]
    max_val = max((r['total'] for r in result), default=1)
    return Response({'data': result, 'max': max_val, 'dias': ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom']})


# ── Registro público de estudiantes ──────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def registro_estudiante(request):
    """Permite crear una cuenta de estudiante sin autenticación previa."""
    data = request.data
    email      = data.get('email', '').strip()
    password   = data.get('password', '')
    first_name = data.get('first_name', '').strip()
    last_name  = data.get('last_name', '').strip()

    if not all([email, password, first_name, last_name]):
        return Response({'error': 'Todos los campos son requeridos'}, status=400)
    if len(password) < 8:
        return Response({'error': 'La contraseña debe tener mínimo 8 caracteres'}, status=400)
    if User.objects.filter(email=email).exists():
        return Response({'error': 'Ya existe una cuenta con ese correo'}, status=400)

    try:
        username = email.split('@')[0] + str(User.objects.count())
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
        )
        user.role         = 'estudiante'
        user.access_point = 'Portal Web'
        user.save()

        # Login automático
        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(user)
        return Response({
            'user':    UserSerializer(user).data,
            'token':   str(refresh.access_token),
            'refresh': str(refresh),
        }, status=201)
    except Exception as e:
        return Response({'error': f'Error al crear cuenta: {str(e)}'}, status=400)